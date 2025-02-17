import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import XLUtils
from openpyxl import load_workbook
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")
file = r"data drivern .xlsx"
workbook = load_workbook(file)
sheet_name = workbook["Sheet1"]
for row in sheet_name.iter_rows(min_row=2,values_only=True):
    print(row)
rows = XLUtils.getRowCount(file,"Sheet1")
cols = XLUtils.getColumnCount(file,"Sheet1")
print(rows, cols)
for row in range(2, rows+1):
    principle = XLUtils.readData(file,"Sheet1",row,1)
    rate_of_interest = XLUtils.readData(file,"Sheet1",row,2)
    tenure = XLUtils.readData(file,"Sheet1",row,3)
    tenure_period = XLUtils.readData(file,"Sheet1",row,4)
    frequency = XLUtils.readData(file, "Sheet1", row, 5)
    expected_maturity = XLUtils.readData(file, "Sheet1", row, 6)

    driver.find_element(By.XPATH,"//input[@name='principal']").clear()
    driver.find_element(By.XPATH,"//input[@name='principal']").send_keys(principle)

    driver.find_element(By.XPATH,"//input[@name='interest']").clear()
    driver.find_element(By.XPATH,"//input[@name='interest']").send_keys(rate_of_interest)

    driver.find_element(By.XPATH,"//input[@name='tenure']").clear()
    driver.find_element(By.XPATH,"//input[@name='tenure']").send_keys(tenure)

    tenure_dropdown = Select(driver.find_element(By.XPATH,"//select[@id='tenurePeriod']"))
    tenure_dropdown.select_by_visible_text(tenure_period)

    frequency_dropdown = Select(driver.find_element(By.XPATH,"//select[@id='frequency']"))
    frequency_dropdown.select_by_visible_text(frequency)

    driver.find_element(By.XPATH, "//a[@href=\"javascript:void(0);\"]").click()

    maturity_value = driver.find_element(By.XPATH,"//span[@id='resp_matval']").text

    if float(maturity_value) == float(expected_maturity):
        XLUtils.writeData(file,"Sheet1",row,8,"Pass")
        XLUtils.fillGreenColor(file,"Sheet1",row,8)
    else:
        XLUtils.writeData(file, "Sheet1", row, 8, "Fail")
        XLUtils.fillRedColor(file, "Sheet1", row, 8)

driver.quit()

# # url = https://www.cit.com/cit-bank/resources/calculators/savings-calculator



# import os
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import Select
# import XLUtils
# from openpyxl import load_workbook
#
# # Initialize WebDriver
# driver = webdriver.Chrome()
# driver.maximize_window()
#
# # Navigate to Fixed Deposit Calculator
# driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")
#
# # Excel File Path
# file = os.path.join(os.getcwd(), "data drivern .xlsx")
# workbook = load_workbook(file)
# sheet_name = workbook["Sheet1"]
#
# # Read Row and Column Count
# rows = XLUtils.getRowCount(file, "Sheet1")
# cols = XLUtils.getColumnCount(file, "Sheet1")
# print(f"Rows: {rows}, Columns: {cols}")
#
# # Iterate through rows
# for row in range(2, rows + 1):
#     # Read data from Excel
#     principle = XLUtils.readData(file, "Sheet1", row, 1)
#     rate_of_interest = XLUtils.readData(file, "Sheet1", row, 2)
#     tenure = XLUtils.readData(file, "Sheet1", row, 3)
#     tenure_period = XLUtils.readData(file, "Sheet1", row, 4)
#     frequency = XLUtils.readData(file, "Sheet1", row, 5)
#     expected_maturity = XLUtils.readData(file, "Sheet1", row, 6)
#
#     # Interact with webpage elements
#     driver.find_element(By.NAME, "principal").clear()
#     driver.find_element(By.NAME, "principal").send_keys(principle)
#
#     driver.find_element(By.NAME, "interest").clear()
#     driver.find_element(By.NAME, "interest").send_keys(rate_of_interest)
#
#     driver.find_element(By.NAME, "tenure").clear()
#     driver.find_element(By.NAME, "tenure").send_keys(tenure)
#
#     # Select dropdown values
#     tenure_dropdown = Select(driver.find_element(By.ID, "tenurePeriod"))
#     tenure_dropdown.select_by_visible_text(tenure_period)
#
#     frequency_dropdown = Select(driver.find_element(By.ID, "frequency"))
#     frequency_dropdown.select_by_visible_text(frequency)
#
#     # Click Calculate
#     driver.find_element(By.XPATH, "//a[contains(text(), 'Calculate')]").click()
#
#     # Fetch and validate maturity value
#     maturity_value = driver.find_element(By.ID, "resp_matval").text
#     maturity_value = maturity_value.replace(",", "")  # Remove commas
#
#     if float(maturity_value) == float(expected_maturity):
#         XLUtils.writeData(file, "Sheet1", row, 7, "Pass")
#         XLUtils.fillGreenColor(file, "Sheet1", row, 7)
#     else:
#         XLUtils.writeData(file, "Sheet1", row, 7, "Fail")
#         XLUtils.fillRedColor(file, "Sheet1", row, 7)
#
# driver.quit()
# print("Test execution completed.")
