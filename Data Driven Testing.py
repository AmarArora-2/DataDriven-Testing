# reading a data

# import openpyxl
# import os
# # file = path
# file = os.getcwd() + r"\Tata_1mg_Test_Scenarios.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Home Page"]
# rows = sheet.max_row
# print("Max rows",rows)
# col = sheet.max_column
# print("Max Columns", col)
# # reading all rows and columns
# for r in range(1,rows+1):
#     for c in range(1,col+1):
#         print(sheet.cell(r,c).value,end="")
#     print()

# ------------------------------------------------------------------

# Writimg a data
# file should be closed

import openpyxl
import os
file = r"/Writeable.xlsx"
# file = os.getcwd() + r"data drivern .xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value='welcome'
workbook.save(file)

# sheet.cell(2,1).value='541'

